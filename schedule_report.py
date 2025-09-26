
import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime,date
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config

st.title("Excel File Reader with Month and Year Filter")

# Corrected the project ID back to the original one from your code
WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"  # Corrected back to the original value
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

if 'processed_df' not in st.session_state:
    st.session_state.processed_df = None
if 'total_count_df' not in st.session_state:
    st.session_state.total_count_df = None
if 'selected_file_name' not in st.session_state:
    st.session_state.selected_file_name = None
if 'sheduledf' not in st.session_state:
    st.session_state.sheduledf = None
if 'shedule' not in st.session_state:
    st.session_state.shedule = None
if 'df_selected' not in st.session_state:
    st.session_state.df_selected = None

def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    
    try:
        response = requests.post(auth_url, headers=headers, data=data)
        if response.status_code != 200:
            st.error(f"Failed to get access token: {response.text}")
            return None
        return response.json()['access_token']
    except Exception as e:
        st.error(f"Error getting access token: {str(e)}")
        return None

def generatePrompt(json_datas):
    # Added fallback mechanism to handle API failures
    try:
        body = {
            "input": f"""
            read this table properly and i need total count of each acitvity as json example provided
            {json_datas}

            Example json needed:
            [{{
                "Activity Name":"name",
                "Total":"Count"
            }}]
    Return only the JSON object, no code, no explanation, just the formatted JSON, and count properly please.
            """, 
            "parameters": {
                "decoding_method": "greedy",
                "max_new_tokens": 8100,
                "min_new_tokens": 0,
                "stop_sequences": [";"],
                "repetition_penalty": 1.05,
                "temperature": 0.5
            },
            "model_id": MODEL_ID,
            "project_id": PROJECT_ID
        }
        
        access_token = GetAccesstoken()
        if not access_token:
            # Create a fallback response by calculating totals ourselves
            return generate_fallback_totals(json_datas)
            
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        response = requests.post(WATSONX_API_URL, headers=headers, json=body)
        
        if response.status_code != 200:
            st.warning(f"WatsonX API failed: {response.text}. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)
            
        return response.json()['results'][0]['generated_text'].strip()
    
    except Exception as e:
        st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
        return generate_fallback_totals(json_datas)


# Function to generate a table of activity names and their Finish dates
def generate_activity_finish_table(df, selected_year, selected_months):
    """
    Generate a table with Activity Names and their Finish dates based on selected year and months.
    
    Parameters:
    - df: Filtered DataFrame containing the processed data
    - selected_year: Integer year selected by the user
    - selected_months: List of month abbreviations (e.g., ['Jan', 'Feb'])
    
    Returns:
    - DataFrame with columns ['Activity Name', 'Finish']
    """
    # Filter the DataFrame for the selected year and months
    filtered_df = df[
        (df['Finish Year'] == selected_year) & 
        (df['Finish Month'].isin(selected_months))
    ]
    
    # Select only the required columns
    result_df = filtered_df[['Activity ID','Activity Name', 'Finish']].copy()
    
    # Sort by Activity Name for consistency
    result_df = result_df.sort_values(by='Activity ID')
    
    # Reset index for clean display
    result_df = result_df.reset_index(drop=True)
    
    return result_df


def generate_fallback_totals(df):
    """Fallback method to calculate totals when the WatsonX API fails"""
    try:
        # Calculate row sums for the DataFrame
        result = []
        for index, row in df.iterrows():
            # Sum all numeric values in the row
            total = row.sum()
            result.append({
                "Activity Name": index,
                "Total": int(total)  # Convert to int to match expected format
            })
        
        # Convert to JSON string
        return json.dumps(result)
    except Exception as e:
        st.error(f"Error in fallback total calculation: {str(e)}")
        # Return an empty JSON array as last resort
        return "[]"
    
def to_excel(df, year, towername, tdatas):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write DataFrame starting from row 3 (Excel row 3, 1-based) to account for new row and title
        df.to_excel(writer, index=True, sheet_name='Activity Counts', startrow=2)
        
        # Get the openpyxl workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Activity Counts']
        
        # Define the title
        title = f"Activity Counts For {towername} Report:({year})"
        
        # Calculate the number of columns (index + DataFrame columns)
        total_columns = len(df.columns) + 1  # +1 for index column
        
        # Add "Tower", "5", "5" in the first row in separate cells (A1, B1, C1)
        worksheet['A1'].value = tdatas['Activity Name'].iloc[0]
        worksheet['B1'].value = tdatas['Start'].iloc[0]
        worksheet['C1'].value = tdatas['Finish'].iloc[0]
        
        # Apply styling to first row cells: bold font, centered
        for col in ['A1', 'B1', 'C1']:
            cell = worksheet[col]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells in the second row for the title
        start_cell = 'A2'
        end_cell = f'{get_column_letter(total_columns)}2'
        worksheet.merge_cells(f'{start_cell}:{end_cell}')
        
        # Set the title text
        title_cell = worksheet['A2']
        title_cell.value = title
        
        # Apply styling to title: yellow background, bold font, centered
        title_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
    return output.getvalue()

def get_cos_files():
    try:
        response = st.session_state.cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            st.warning("No .xlsx files found in the bucket 'projectreport'. Please ensure Excel files are uploaded.")
        return files
    except Exception as e:
        st.error(f"Error fetching COS files Please Check The Internet Connection")
        return ["Error Fetching Files"]

def getTotal(ai_data):
    try:
        # Handle case where ai_data might be a string
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
            
        share = []
        for i in ai_data:
            share.append(i['Total'])
        return share
    except Exception as e:
        st.error(f"Error parsing AI data: {str(e)}")
        return [0] * len(st.session_state.sheduledf.index)  # Return zeros as fallback

# Improved function to check if a cell's font is bold
def is_cell_bold(cell):
    # Check if font exists and bold property is True
    return cell.font and getattr(cell.font, 'bold', None) is True

# Function to process the Excel file - improved with consistent bold detection
def process_file(file_stream):
    workbook = pd.ExcelFile(file_stream)
    
    if "TOWER 5 FINISHING." in workbook.sheet_names:
        sheet_name = "TOWER 5 FINISHING."
        
        # Read the data with header at row 1 (Excel row 1, 0-based index 0)
        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=0)
        
        st.session_state.date = df[['Activity Name', 'Start', 'Finish']].head().iloc[1:2]
        # Assign column names based on document structure
        expected_columns = [
            'Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
            'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
            'Actual Start', 'Actual Finish', '% Complete', 'Start', 'Finish', 'Delay Reasons'
        ]
        
        if len(df.columns) >= len(expected_columns):
            df.columns = expected_columns[:len(df.columns)]
        else:
            st.error("Excel file has fewer columns than expected.")
            return None, None
        
        # Select desired columns
        target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Start", "Finish"]
        df = df[target_columns]
        
        # List of activity names to filter
        activity_names = [
            "Brickwork",
            "AC Installation",
            "Balconies Waterproofing",
            "Brick masonry for entrance wall",
            "C-F-First Fix",
            "C-Gypsum and POP Punning",
            "C-P-First Fix",
            "C-Stone flooring",
            "Closing of shafts",
            "Copper Piping",
            "Counter stone works",
            "CP-Final Fix",
            "EL-Final Fix",
            "EL-Second Fix",
            "False ceiling framing",
            "Fixing of brackets for GRC Moduling",
            "Floor Tiling",
            "Glass Installation",
            "GRC jali fixing (Fire escape staicase)",
            "GRC jali fixing (main staircase)",
            "GRC jali fixing (splash pool)",
            "GRC molding fixing",
            "Grouting of toilets & balcony Tiles",
            "Gypsum board false ceiling",
            "Installation of Rear & Front balcony UPVC Windows",
            "Installation of doors",
            "Installation of wardrobes and cabinets",
            "Ledge Wall Construction",
            "MS works in balconies",
            "Paint in balcony and shafts",
            "Painting First Coat",
            "SS Framing",
            "ST-Electrical",
            "ST-Fire fighting",
            "ST-Plumbing & Water supply",
            "Stone cills, ledges and jambs",
            "Texture paint (final coat)",
            "Texture paint (first coat)",
            "Wall Tiling",
            "Water Proofing Works",
            "Waterproofing works"
        ]
        
        # Filter rows where Activity Name matches one of the specified activity_names
        df_filtered = df[df['Activity Name'].isin(activity_names)]
        
        # Convert 'Finish' column to datetime
        df_filtered['Finish'] = pd.to_datetime(df_filtered['Finish'], errors='coerce')
        
        # Extract month and year from the 'Finish' column
        df_filtered['Finish Month'] = df_filtered['Finish'].dt.strftime('%b')
        df_filtered['Finish Year'] = df_filtered['Finish'].dt.year
        
        return df_filtered, "Tower 5"
    
    elif "TOWER 4 FINISHING." in workbook.sheetnames:
        sheet_name = "TOWER 4 FINISHING."
        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=0)
        
        st.session_state.date = df[['Activity Name', 'Start', 'Finish']].head().iloc[1:2]

        # Assign column names
        expected_columns = [
            'Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
            'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
            'Actual Start', 'Actual Finish', '% Complete', 'Start', 'Finish', 'Delay Reasons'
        ]
        
        if len(df.columns) >= len(expected_columns):
            df.columns = expected_columns[:len(df.columns)]
        else:
            st.error("Excel file has fewer columns than expected.")
            return None, None
        
        # Select desired columns
        target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Start", "Finish"]
        df = df[target_columns]
        
        # Define the column index for 'Activity Name' (0-based index, column F = index 5)
        activity_col_idx = 5
        
        # Get row indices where Activity Name is not bold, starting from Excel row 2
        non_bold_rows = [
            row_idx for row_idx, row in enumerate(workbook[sheet_name].iter_rows(min_row=2, max_col=16), start=0)
            if row[activity_col_idx].value and (not row[activity_col_idx].font or not row[activity_col_idx].font.b)
        ]
        
        # Extract rows where Activity Name is not bold
        if non_bold_rows:
            df_non_bold = df.iloc[non_bold_rows]
        else:
            df_non_bold = pd.DataFrame(columns=df.columns)
        
        # Convert 'Finish' column to datetime
        df_non_bold['Finish'] = pd.to_datetime(df_non_bold['Finish'], errors='coerce')
        
        # Extract Month Name and Year
        df_non_bold['Finish Month'] = df_non_bold['Finish'].dt.strftime('%b')
        df_non_bold['Finish Year'] = df_non_bold['Finish'].dt.year
        
        return df_non_bold, "Tower 4"
    
    else:
        return None, None
    
def StoreExcelFile(excel_data, file_name):
    api_key = "jocRH1td-ZjJSqlFwFtjjTRzUDgDgTaa6gM99nr1W-e6"
    instanceid = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:schedulereport1"
    endpoint = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
    cosbucket = "schedulereport1"

    cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=api_key,
    ibm_service_instance_id=instanceid,
    config=Config(signature_version='oauth'),
    endpoint_url=endpoint
)
    # cos_client.upload_fileobj(excel_data, cosbucket, "test.xlsx")
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        excel_data.to_excel(writer, index=False)
    output.seek(0)  # Reset buffer position to the start

    # Upload the file-like object
    cos_client.upload_fileobj(output, cosbucket, file_name)

    
# Streamlit App
st.title("Excel File Activity Processor")

# Initialize session state
if 'df_selected' not in st.session_state:
    st.session_state.df_selected = None
if 'tname' not in st.session_state:
    st.session_state.tname = None
if 'selected_file' not in st.session_state:
    st.session_state.selected_file = None

# Get files from COS
files = get_cos_files()

st.sidebar.header("Select a File")
selected_file = st.sidebar.selectbox("Choose a file from IBM COS:", files, key="file_selector")

# Process file only if a new file is selected
if selected_file and selected_file != st.session_state.selected_file:
    st.session_state.selected_file = selected_file
    try:
        response = st.session_state.cos_client.get_object(Bucket="projectreportnew", Key=selected_file)
        file_bytes = io.BytesIO(response['Body'].read())
        st.session_state.df_selected, st.session_state.tname = process_file(file_bytes)
    except Exception as e:
        # st.error(f"Error loading file from cloud storage: {str(e)}")
        st.session_state.df_selected = None
        st.session_state.tname = None



# Work with the stored DataFrame
df_selected = st.session_state.df_selected
tname = st.session_state.tname

if df_selected is None:
    st.warning("No valid data found in the selected file. Ensure the file contains 'TOWER 5 FINISHING.' or 'TOWER 4 FINISHING.' sheet.")
else:
    st.write("Processed Data:")
    st.write(df_selected)   

    # Check if there's data to filter
    if not df_selected.empty:
        # Get unique years and months for filters
        unique_years = sorted(df_selected['Finish Year'].dropna().unique().astype(int))
        unique_months = sorted(df_selected['Finish Month'].dropna().unique())

        # Only show filters if we have data
        if unique_years and unique_months:
            # Sidebar filters
            st.sidebar.header("Filters")
            selected_year = st.sidebar.selectbox("Select Year", unique_years, key="year_filter")
            selected_months = st.sidebar.multiselect("Select Months", unique_months, default=unique_months, key="month_filter")

            # Apply filters directly on the stored DataFrame
            filtered_data = df_selected[
                (df_selected['Finish Year'] == selected_year) & 
                (df_selected['Finish Month'].isin(selected_months))
            ]

            st.subheader(f"Filtered Data for {', '.join(selected_months)} {selected_year}")
            st.write(filtered_data)

            # Generate and display the Activity Names and Finish Dates table
            # st.subheader(f"Activity Names and Finish Dates for {', '.join(selected_months)} {selected_year}")
            activity_finish_table = generate_activity_finish_table(df_selected, selected_year, selected_months)
            
            if not activity_finish_table.empty:
                st.write(activity_finish_table)
            else:
                st.warning(f"No activities found for {', '.join(selected_months)} {selected_year}.")

            # Debug: Count of specific activity for verification
            if 'Balconies Waterproofing' in filtered_data['Activity Name'].values:
                count_running = filtered_data[filtered_data['Activity Name'] == 'Balconies Waterproofing'].shape[0]
                st.write(f"Count of 'Balconies Waterproofing' activities: {count_running}")

            if st.button("Display Activity Count by Month"):
                try:
                    st.write(st.session_state.date)
                    # StoreExcelFile(activity_finish_table)
                    st.write(tname)
                    today = date.today()
                    # StoreExcelFile(activity_finish_table,f"activity_names{', '.join(selected_months)}_{selected_year} {tname} ({today}).xlsx")
                    st.session_state.timed = f"activity_names{', '.join(selected_months)}_{selected_year} {tname}_({today}).xlsx"
                    st.write(st.session_state.timed)
                    StoreExcelFile(activity_finish_table,f"activity_names{', '.join(selected_months)}_{selected_year} {tname}_({today}).xlsx")
                    # Get only necessary columns for activity and month
                    activity_month_data = filtered_data[['Activity Name', 'Finish Month']]

                    # Group and count activities by month
                    count_table = (
                        activity_month_data
                        .groupby(['Activity Name', 'Finish Month'])
                        .size()
                        .reset_index(name='Count')
                        .pivot(index='Activity Name', columns='Finish Month', values='Count')
                        .fillna(0)
                        .astype(int)
                    )

                    # Sort months in calendar order
                    month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                    
                    # Only include months that exist in the data
                    available_months = [m for m in month_order if m in count_table.columns]
                    if available_months:
                        count_table = count_table.reindex(columns=available_months)

                    # Store the count table for safe access in the getTotal function
                    st.session_state.sheduledf = count_table

                    # WatsonX API call with try-except and fallback
                    try:
                        test = generatePrompt(count_table)
                        ai_data = json.loads(test)
                    except json.JSONDecodeError:
                        st.warning("Failed to get AI-generated totals. Using fallback method.")
                        # Generate a fallback response directly
                        ai_data = []
                        for idx, row in count_table.iterrows():
                            ai_data.append({
                                "Activity Name": idx,
                                "Total": int(row.sum())
                            })

                    # Add the Total column
                    count_table['Total'] = getTotal(ai_data)

                    st.write("Activity Count by Month:")
                    st.write(count_table)
                    st.session_state.sheduledf = count_table
                    current_date = datetime.now().date()
                    st.session_state.shedule = to_excel(count_table, current_date, tname, st.session_state.date)
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=st.session_state.shedule,
                        file_name=f"Shedule_Report_{tname}_{selected_year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Error generating activity count table: {str(e)}")
        else:
            st.warning("No date information available in the processed data.")
    else:
        st.warning("The processed data is empty. No rows were selected based on the bold text criteria.")
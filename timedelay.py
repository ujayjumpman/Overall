import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl import Workbook


time_delays_days = []



WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"  # Corrected back to the original value
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"



def current_tracker_cos():
    COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
    COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
    COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
    COS_BUCKET = "projectreportnew"
    st.session_state.cos_client = ibm_boto3.client(
        's3',
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
        config=Config(signature_version='oauth'),
        endpoint_url=COS_ENDPOINT
    )
    try:
        response = st.session_state.cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            st.warning("No .xlsx files found in the bucket 'projectreport'. Please ensure Excel files are uploaded.")
        return files
    except Exception as e:
        st.error(f"Error fetching COS files: {str(e)}")
        return []
    
def stored_tracker_cos():
    COS_API_KEY = "axCN_xatDLCDPi1YNw6WTtzefxNPoX9-2csNGUoByv3f"
    COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:schedulereport1"
    COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
    COS_BUCKET = "schedulereport1"
    st.session_state.cos_client = ibm_boto3.client(
        's3',
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
        config=Config(signature_version='oauth'),
        endpoint_url=COS_ENDPOINT
    )
    try:
        response = st.session_state.cos_client.list_objects_v2(Bucket=COS_BUCKET)
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            st.warning("No .xlsx files found in the bucket 'projectreport'. Please ensure Excel files are uploaded.")
        return files
    except Exception as e:
        st.error(f"Error fetching COS files: {str(e)}")
        return []



def process_file(filestream, workbook):
    if "TOWER 5 FINISHING." in workbook.sheet_names:
        df = pd.read_excel(
            filestream,
            sheet_name="TOWER 5 FINISHING.",
            engine="openpyxl",
            header=0
        )
        st.session_state.towername = "Tower 5"
        activity_names = [
            "Brickwork", "AC Installation", "Balconies Waterproofing", "Brick masonry for entrance wall",
            "C-F-First Fix", "C-Gypsum and POP Punning", "C-P-First Fix", "C-Stone flooring",
            "Closing of shafts", "Copper Piping", "Counter stone works", "CP-Final Fix",
            "EL-Final Fix", "EL-Second Fix", "False ceiling framing", "Fixing of brackets for GRC Moduling",
            "Floor Tiling", "Glass Installation", "GRC jali fixing (Fire escape staicase)",
            "GRC jali fixing (main staircase)", "GRC jali fixing (splash pool)", "GRC molding fixing",
            "Grouting of toilets & balcony Tiles", "Gypsum board false ceiling",
            "Installation of Rear & Front balcony UPVC Windows", "Installation of doors",
            "Installation of wardrobes and cabinets", "Ledge Wall Construction", "MS works in balconies",
            "Paint in balcony and shafts", "Painting First Coat", "SS Framing", "ST-Electrical",
            "ST-Fire fighting", "ST-Plumbing & Water supply", "Stone cills, ledges and jambs",
            "Texture paint (final coat)", "Texture paint (first coat)", "Wall Tiling",
            "Water Proofing Works", "Waterproofing works"
        ]
        df['Finish'] = pd.to_datetime(df['Finish'])
        df['Finish Month'] = df['Finish'].dt.strftime('%b')
        df['Finish Year'] = df['Finish'].dt.year
        filtered_df = df[df['Activity Name'].isin(activity_names)]
        return filtered_df[['Activity ID', 'Activity Name', '% Complete', 'Finish', 'Finish Month', 'Finish Year']]
    
    elif "TOWER 4 FINISHING." in workbook.sheet_names:
        sheet_name = "TOWER 4 FINISHING."
        df = pd.read_excel(filestream, sheet_name=sheet_name, header=0)
        
        st.session_state.date = df[['Activity Name', 'Start', 'Finish']].head().iloc[1:2] 
        st.session_state.towername = "Tower 4"
        
        expected_columns = [
            'Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
            'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
            'Actual Start', 'Actual Finish', '% Complete', 'Start', 'Finish', 'Delay Reasons'
        ]
        
        if len(df.columns) >= len(expected_columns):
            df.columns = expected_columns[:len(df.columns)]
        else:
            st.error("Excel file has fewer columns than expected.")
            return None
        
        target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Start", "Finish"]
        df = df[target_columns]
        
        activity_col_idx = 5
        non_bold_rows = [
            row_idx for row_idx, row in enumerate(workbook[sheet_name].iter_rows(min_row=2, max_col=16), start=0)
            if row[activity_col_idx].value and (not row[activity_col_idx].font or not row[activity_col_idx].font.b)
        ]
        
        if non_bold_rows:
            df_non_bold = df.iloc[non_bold_rows]
        else:
            df_non_bold = pd.DataFrame(columns=df.columns)
        
        df_non_bold['Finish'] = pd.to_datetime(df_non_bold['Finish'], errors='coerce')
        df_non_bold['Finish Month'] = df_non_bold['Finish'].dt.strftime('%b')
        df_non_bold['Finish Year'] = df_non_bold['Finish'].dt.year
        
        return df_non_bold
    
    else:
        st.write("No data available")
        return None

# Existing process_for_stored_tracker function (unchanged)
def process_for_stored_tracker(filestream):
    df = pd.read_excel(filestream)
    df['Finish'] = pd.to_datetime(df['Finish'])
    df['Finish Month'] = df['Finish'].dt.strftime('%b')
    df['Finish Year'] = df['Finish'].dt.year
    return df

# New function to compute days difference between current and stored tracker Finish dates
def compute_finish_days_difference(current_df, stored_df):
    if current_df is not None and stored_df is not None and not current_df.empty and not stored_df.empty:
        # Ensure 'Finish' columns are in datetime format
        current_df['Finish'] = pd.to_datetime(current_df['Finish'], errors='coerce')
        stored_df['Finish'] = pd.to_datetime(stored_df['Finish'], errors='coerce')
        
        # Merge DataFrames on 'Activity ID' (inner join to include only matching IDs)
        merged_df = pd.merge(
            current_df[['Activity ID', 'Activity Name','Finish']],
            stored_df[['Activity ID', 'Finish']],
            on='Activity ID',
            how='inner',
            suffixes=('_current', '_stored')
        )
        
        # Calculate days difference (current Finish - stored Finish)
        merged_df['Days Difference'] = (merged_df['Finish_current'] - merged_df['Finish_stored']).dt.days
        
        # Select relevant columns
        result_df = merged_df[['Activity ID', 'Activity Name','Days Difference']]
            
        return result_df
    return pd.DataFrame(columns=['Activity ID', 'Activity Name', 'Days Difference'])


def find_max_positive_repeated_activities(days_diff_df):
    if days_diff_df is not None and not days_diff_df.empty:
        # Filter for positive Days Difference
        positive_df = days_diff_df[days_diff_df['Days Difference'] > 0]
        
        if positive_df.empty:
            return pd.DataFrame(columns=['SNo', 'Activity Name', 'DelayDays', 'DelayReasons', 'Remarks'])
        
        # Find Activity Names that are repeated (appear more than once)
        activity_counts = positive_df['Activity Name'].value_counts()
        repeated_activities = activity_counts[activity_counts > 1].index
        
        if not repeated_activities.empty:
            # Filter for repeated Activity Names
            repeated_df = positive_df[positive_df['Activity Name'].isin(repeated_activities)]
            
            # Group by Activity Name and find the maximum Days Difference
            max_values = repeated_df.groupby('Activity Name')['Days Difference'].max().reset_index()
            
            # Merge to get rows where Days Difference equals the maximum, then drop duplicates
            result_df = pd.merge(
                repeated_df,
                max_values,
                on=['Activity Name', 'Days Difference'],
                how='inner'
            )
            
            # Drop duplicates to keep only the first occurrence of each Activity Name
            result_df = result_df.drop_duplicates(subset=['Activity Name'], keep='first')
            
            # Sort by Activity Name and Days Difference
            result_df = result_df.sort_values(by=['Activity Name', 'Days Difference'], ascending=[True, False])
            
            # Create output DataFrame
            output_df = pd.DataFrame({
                'SNo': range(1, len(result_df) + 1),
                'Tower Name': st.session_state.towername,
                'Activity Name': result_df['Activity Name'],
                'DelayDays': result_df['Days Difference'],
                'DelayReasons': '',
                'Remarks': ''
            })
            
            return output_df
        
    return pd.DataFrame(columns=['SNo', 'Activity Name', 'DelayDays', 'DelayReasons', 'Remarks'])

#ithu vanthu excel download panrathukaana code
def create_excel_file(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Repeated Activities"
    
    # Add title in the first row
    title = f"Time Delay Report({datetime.now().strftime('%Y-%m-%d')})"
    ws.cell(row=1, column=1).value = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)  # Merge across 6 columns now
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    
    # Define headers in the second row
    headers = ['SNo', 'Tower', 'Activity Name', 'Delay Days', 'Delay Reason', 'Remarks']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Add data starting from the third row
    for row_idx, row in enumerate(df.values, start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Adjust column widths for better readability
    column_widths = {'A': 8, 'B': 20, 'C': 12, 'D': 15, 'E': 30, 'F': 30}  # Updated for new column
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    
    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Main logic
unique_years = []
if stored_tracker := st.sidebar.selectbox("Choose a Stored Tracker File:", stored_tracker_cos(), key="stored_tracker_file_selector"):
    response = st.session_state.cos_client.get_object(Bucket="schedulereport1", Key=stored_tracker)
    file_stream = io.BytesIO(response['Body'].read())
    st.session_state.stored_tracker = process_for_stored_tracker(file_stream)
    if st.session_state.stored_tracker is not None:
        unique_years = st.session_state.stored_tracker['Finish Year'].dropna().unique().astype(int)
        unique_years = sorted(unique_years)

st.sidebar.header("Select a File")
current_tracker = st.sidebar.selectbox("Choose a Current Tracker File:", current_tracker_cos(), key="current_tracker_file_selector")

st.sidebar.header("Filter by Month and Year")
month_filter = st.sidebar.multiselect("Choose Month(s):", ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], key="month_filter")
year_filter = st.sidebar.multiselect("Choose Year(s):", unique_years if unique_years else list(range(2020, 2026)), key="year_filter")

# Process and display current tracker
if current_tracker:
    try:
        # Retrieve file from IBM COS
        response = st.session_state.cos_client.get_object(Bucket="projectreportnew", Key=current_tracker)
        file_stream = io.BytesIO(response['Body'].read())
        
        # Process Excel file
        try:
            workbook = pd.ExcelFile(file_stream, engine="openpyxl")
            st.session_state.current_tracker = process_file(file_stream, workbook)
            
            if st.session_state.current_tracker is not None:
                filtered_current_df = st.session_state.current_tracker
                
                # Apply month and year filters
                try:
                    if month_filter:
                        if 'Finish Month' not in filtered_current_df.columns:
                            st.error("Error: 'Finish Month' column missing in Current Tracker.")
                        else:
                            filtered_current_df = filtered_current_df[filtered_current_df['Finish Month'].isin(month_filter)]
                    if year_filter:
                        if 'Finish Year' not in filtered_current_df.columns:
                            st.error("Error: 'Finish Year' column missing in Current Tracker.")
                        else:
                            filtered_current_df = filtered_current_df[filtered_current_df['Finish Year'].isin(year_filter)]
                    
                    st.write("Current Tracker Data:")
                    st.write(filtered_current_df)
                except Exception as e:
                    st.error(f"Error filtering Current Tracker data: {str(e)}")
            else:
                st.error("Please Choose a Valid Current Tracker File")
        except Exception as e:
            st.error(f"Error reading Current Tracker Excel file: {str(e)}")
    except Exception as e:
        st.error(f"Unexpected error with Current Tracker: {str(e)}")
else:
    st.warning("Please select a Current Tracker file.")

# Process and display stored tracker (unchanged)
if stored_tracker:
    try:
        filtered_stored_df = st.session_state.stored_tracker
        
        try:
            if month_filter:
                if 'Finish Month' not in filtered_stored_df.columns:
                    st.error("Error: 'Finish Month' column missing in Stored Tracker.")
                else:
                    filtered_stored_df = filtered_stored_df[filtered_stored_df['Finish Month'].isin(month_filter)]
            if year_filter:
                if 'Finish Year' not in filtered_stored_df.columns:
                    st.error("Error: 'Finish Year' column missing in Stored Tracker.")
                else:
                    filtered_stored_df = filtered_stored_df[filtered_stored_df['Finish Year'].isin(year_filter)]
            
            st.write("Stored Tracker Data:")
            st.write(filtered_stored_df)
        except Exception as e:
            st.error(f"Error filtering Stored Tracker data: {str(e)}")
    except Exception as e:
        st.error(f"Error processing Stored Tracker: {str(e)}")
else:
    st.warning("Please select a Stored Tracker file.")

# Compute and display days difference between Finish dates
if current_tracker and stored_tracker:
    try:
        days_diff_df = compute_finish_days_difference(filtered_current_df, filtered_stored_df)
        if not days_diff_df.empty:
            st.write("Days Difference (Current Tracker Finish - Stored Tracker Finish):")
            st.write(days_diff_df)
        else:
            st.warning("No matching Activity IDs found between Current and Stored Trackers.")
    except Exception as e:
        st.error(f"Please Choose a Valid Current Tracker File")
else:
    st.warning("Please select both Current and Stored Tracker files to compute days difference.")

# Compute and display max positive repeated activities with downloadable Excel
if current_tracker and stored_tracker:
    try:
        days_df = find_max_positive_repeated_activities(days_diff_df)
        if not days_df.empty:
            st.write("Max Positive Repeated Activities:")
            st.write(days_df)
            
            # Create and provide downloadable Excel file
            try:
                excel_buffer = create_excel_file(days_df)
                st.download_button(
                    label="Time Delay Report",
                    data=excel_buffer,
                    file_name="repeated_activities_delay_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_repeated_activities"
                )
            except Exception as e:
                st.error(f"Error creating Excel file for download: {str(e)}")
        else:
            st.warning("No repeated Activity Names with positive Delay Days found.")
    except Exception as e:
        st.error(f"Please Choose a Valid Current Tracker File")
else:
    st.warning("Please select both Current and Stored Tracker files to compute repeated activities.")
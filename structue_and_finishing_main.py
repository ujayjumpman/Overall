import streamlit as st
from structure_and_finishing1 import *
from structure_and_finishing2 import *
from structure_and_finishing3 import *
from structure_and_finishing4 import *
from io import BytesIO
import pandas as pd
import urllib.parse
from datetime import datetime
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter



def get_cos_files():
    try:
        response = st.session_state.cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return []
    

st.title("Multiple Excel File Sheet Name Viewer")

# Convert to Excel in memory
def to_excel(df):
    try:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Combined')
        processed_data = output.getvalue()
        return processed_data
    except Exception as e:
        st.error(f"Error generating Excel file: {str(e)}")
        return None

# uploaded_files = st.file_uploader("Upload Excel files", type=["xlsx", "xls"], accept_multiple_files=True)

# if uploaded_files:
#     averagedf = None
#     averagedf2 = None
#     averagedf3 = None
#     veridia = pd.DataFrame(Getprecentage(uploaded_files))
#     for file in uploaded_files:
#         if "Structure Work Tracker EWS LIG P4 all towers" in file.name:
#             st.write("Processing first file")
#             averagedf = CountingProcess(file)
#         elif "Structure Work Tracker Tower G & Tower H" in file.name:
#             st.write("Processing second file")
#             averagedf2 = CountingProcess2(file)
#         elif "Structure Work Tracker Tower 6 & Tower 7" in file.name:
#             st.write("Processing third file")
#             averagedf3 = CountingProcess3(file)

files = get_cos_files()    

averagedf = None
averagedf2 = None
averagedf3 = None



test = []
veridia = Getprecentage(files)
# veridia = pd.DataFrame(a)


for file in files:
    if "Structure Work Tracker EWS LIG P4 all towers" in file:
        st.write("One")
        response = st.session_state.cos_client.get_object(Bucket="projectreportnew", Key=file)
        st.write("Processing first file")
        averagedf = CountingProcess(io.BytesIO(response['Body'].read()))
    elif "Structure Work Tracker Tower G & Tower H" in file:
        st.write("Two")
        st.write("Processing second file")
        response = st.session_state.cos_client.get_object(Bucket="projectreportnew", Key=file)
        averagedf2 = CountingProcess2(io.BytesIO(response['Body'].read()))
    elif "Structure Work Tracker Tower 6 & Tower 7" in file:
        st.write("Three")
        st.write("Processing third file")
        response = st.session_state.cos_client.get_object(Bucket="projectreportnew", Key=file)
        averagedf3 = CountingProcess3(io.BytesIO(response['Body'].read()))
        

# st.write(veridia)

veridia_df = pd.DataFrame(veridia)
st.write(veridia_df)
if averagedf is not None:
    st.write(averagedf)
    averagedf_df = pd.DataFrame(averagedf)
    st.write(averagedf_df)
if averagedf2 is not None:
    # st.write(averagedf2)
    averagedf2_df = pd.DataFrame(averagedf2)
    st.write(averagedf2_df)
if averagedf3 is not None:
    st.write(averagedf3)
    averagedf3_df = pd.DataFrame(averagedf3)
    st.write(averagedf3_df)

combined_df = pd.concat([veridia_df, averagedf_df, averagedf2_df, averagedf3_df], ignore_index=True)
st.write(combined_df)

st.session_state.overalldf = combined_df

# Define the title
title = f"OVERALL REPORT {datetime.now().date()}"

with io.BytesIO() as buffer:
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Write DataFrame starting from row 2 (to leave space for the title)
        combined_df.to_excel(writer, sheet_name="Combined Data", startrow=1, index=False)
        
        # Get the openpyxl workbook and worksheet
        workbook = writer.book
        worksheet = workbook["Combined Data"]
        
        # Calculate the number of columns in the DataFrame
        total_columns = len(combined_df.columns)
        
        # Merge cells in the first row for the title
        start_cell = 'A1'
        end_cell = f'{get_column_letter(total_columns)}1'
        worksheet.merge_cells(f'{start_cell}:{end_cell}')
        
        # Set the title text
        title_cell = worksheet['A1']
        title_cell.value = title
        
        # Apply styling: yellow background, bold font, centered
        title_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
    buffer.seek(0)
    st.session_state.overall = buffer.getvalue()
    st.download_button(
        label="Download Combined Excel File",
        data=buffer,
        file_name="combined_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )    # test.append(veridia)
    # test.append(averagedf)
    # test.append(averagedf2)
    # test.append(averagedf3)
    # st.write(test)

    # Combine DataFrames
    # combined_dfs = [df for df in [veridia, averagedf, averagedf2, averagedf3] if df is not None and not df.empty]
    # if combined_dfs:
    #     combined_df = pd.concat(combined_dfs, ignore_index=True)
    #     st.write("### 🧾 Combined Data", combined_df)
    #     # st.json(combined_df)
    #     updated_value = update_finishing_value(combined_df)
    #     st.session_state.structure_and_finishingdf = updated_value
    #     st.session_state.structure_and_finishing = to_excel(updated_value)
    #     if excel_data:
    #         st.download_button(
    #             label="📥 Download Combined Excel",
    #             data=excel_data,
    #             file_name="Combined_Tracker_Data.xlsx",
    #             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #         )
    # else:
    #     st.warning("No data to combine.")
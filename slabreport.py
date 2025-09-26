import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from io import BytesIO
from datetime import datetime
import requests
import json
from io import BytesIO
import ibm_boto3
from ibm_botocore.client import Config
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill


COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"



st.session_state.cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)



def get_cos_files():
    try:
        response = st.session_state.cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return []

all_tower_tables = []

WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"

def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']
    


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        i need a total of completed and noncompleted of each months and return as json 

        Note: for json completed and non-completed is inly enough dont add any month name

        Sample json:
        {{
          "completed":"total",
          "non-completed":"total"
        }}

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()




def get_cell_color(workbook, sheet_name, row_num, col_letter):
    sheet = workbook[sheet_name]
    col_index = column_index_from_string(col_letter.upper())
    cell = sheet.cell(row=row_num, column=col_index)
    
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        color_obj = fill.start_color
        if color_obj.type == 'rgb' and color_obj.rgb:
            return f"#{color_obj.rgb[-6:]}", cell.value
        elif color_obj.type == 'theme':
            return "Theme color (not directly readable)", cell.value
        else:
            return "Unknown color format", cell.value
    return "No Fill", cell.value

def get_unique_years(wb, sheet_name):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    all_cols = [
        ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P'],  # Tower2
        ['T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH'],  # Tower3
        ['AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX', 'AZ', 'BB', 'BD', 'BF', 'BH', 'BJ', 'BL', 'BN', 'BP'],  # Tower4
        ['DC', 'DE', 'DG', 'DI', 'DK', 'DM', 'DO', 'DQ', 'DS', 'DU', 'DW', 'DY', 'EA', 'EC'],  # Tower5
        ['FK', 'FM', 'FO', 'FQ', 'FS', 'FU', 'FW', 'FY', 'GA', 'GB', 'GC', 'GE', 'GG', 'GI', 'GK'],  # Tower6
        ['EG', 'EL', 'EK', 'EM', 'EO', 'EQ', 'ES', 'EU', 'EW', 'EY', 'FA', 'FC', 'FE', 'FG']  # Tower7
    ]
    years = set()
    
    for cols in all_cols:
        for i in cols:
            for j in rows:
                _, val = get_cell_color(wb, sheet_name, j, i)
                try:
                    if isinstance(val, datetime):
                        years.add(val.year)
                    else:
                        cell_date = pd.to_datetime(val, errors='coerce')
                        if not pd.isna(cell_date):
                            years.add(cell_date.year)
                except:
                    continue
    
    return sorted(list(years)) if years else [datetime.now().year]

st.title("📊 Excel Cell Color Viewer")

# Initialize lists for each tower
tower2 = []
tower3 = []
tower4 = []
tower5 = []
tower6 = []
tower7 = []

# Sidebar for month and year filters
st.sidebar.header("Date Filters")
months = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12
}
month_names = {v: k[:3].upper() for k, v in months.items()}  # JAN, FEB, etc.
selected_months = st.sidebar.multiselect(
    "Select Months",
    options=list(months.keys()),
    default=list(months.keys())  # Default to all months
)
selected_month_nums = [months[month] for month in selected_months]

def process_tower(tower_name, cols, rows, tower_list, wb, sheet_name, selected_months, selected_year):
    st.write(f"{tower_name}..")
    tower_list.clear()  # Clear the list to avoid duplicate entries
    # Initialize counts for each selected month
    month_counts = {month: {"green": 0, "blue": 0} for month in selected_months}
    
    for i in cols:
        for j in rows:
            color, val = get_cell_color(wb, sheet_name, j, i)
            # Check if the cell value is a date and matches the selected months and year
            try:
                if isinstance(val, datetime):
                    cell_month = val.month
                    cell_year = val.year
                else:
                    # Attempt to parse the value as a date
                    cell_date = pd.to_datetime(val, errors='coerce')
                    if pd.isna(cell_date):
                        continue  # Skip if the value cannot be parsed as a date
                    cell_month = cell_date.month
                    cell_year = cell_date.year
                
                # Only process if the cell's month and year match the selected filters
                if cell_month in selected_months and cell_year == selected_year:
                    if color == '#92D050':
                        tower_list.append(1)
                        month_counts[cell_month]["green"] += 1
                        # st.write(f"Row {j}, Col {i}: Color {color}, Value {val}")
                    elif color == '#00B0F0':
                        tower_list.append(0)
                        month_counts[cell_month]["blue"] += 1
                        # st.write(f"Row {j}, Col {i}: Color {color}, Value {val}")
            except Exception as e:
                st.warning(f"Error processing cell {i}{j}: {e}")
                continue
    
    # Prepare data for table
    table_data = [
        {"Category": "Completed", **{month_names[month]: month_counts[month]["green"] for month in selected_months}},
        {"Category": "Non-Completed", **{month_names[month]: month_counts[month]["blue"] for month in selected_months}}
    ]
    json_totals = json.loads(generatePrompt(pd.DataFrame(table_data)))


    for row in table_data:
        if row["Category"] == "Completed":
            row["Total"] = json_totals["completed"]
        else:
            row["Total"] = json_totals["non-completed"]

    all_tower_tables.append({"tower_name": tower_name, "table_data": table_data})

    # Display table
    if table_data and selected_months:
        st.write(f"{tower_name} Counts by Month")
        st.table(pd.DataFrame(table_data))
    else:
        st.write(f"No data found for {tower_name} in the selected months and year.")

def Tower2(wb, sheet_name, selected_months, selected_year):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    col = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    process_tower("Tower2", col, rows, tower2, wb, sheet_name, selected_months, selected_year)

def Tower3(wb, sheet_name, selected_months, selected_year):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    col = ['T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH']
    process_tower("Tower3", col, rows, tower3, wb, sheet_name, selected_months, selected_year)

def Tower4(wb, sheet_name, selected_months, selected_year):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    col = ['AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX', 'AZ', 'BB', 'BD', 'BF', 'BH', 'BJ', 'BL', 'BN', 'BP']
    process_tower("Tower4", col, rows, tower4, wb, sheet_name, selected_months, selected_year)

def Tower5(wb, sheet_name, selected_months, selected_year):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    col = ['DC', 'DE', 'DG', 'DI', 'DK', 'DM', 'DO', 'DQ', 'DS', 'DU', 'DW', 'DY', 'EA', 'EC']
    process_tower("Tower5", col, rows, tower5, wb, sheet_name, selected_months, selected_year)

def Tower6(wb, sheet_name, selected_months, selected_year):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    col = ['FK', 'FM', 'FO', 'FQ', 'FS', 'FU', 'FW', 'FY', 'GA', 'GB', 'GC', 'GE', 'GG', 'GI', 'GK']
    process_tower("Tower6", col, rows, tower6, wb, sheet_name, selected_months, selected_year)

def Tower7(wb, sheet_name, selected_months, selected_year):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    col = ['EG', 'EL', 'EK', 'EM', 'EO', 'EQ', 'ES', 'EU', 'EW', 'EY', 'FA', 'FC', 'FE', 'FG']
    process_tower("Tower7", col, rows, tower7, wb, sheet_name, selected_months, selected_year)



def create_excel_file(tables, selected_year):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        current_row = 0
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for table in tables:
            tower_name = table["tower_name"]
            df = pd.DataFrame(table["table_data"])
            
            # Write tower name as header
            pd.DataFrame([[f"{tower_name} Counts by Month (Year: {selected_year})"]], columns=[""]).to_excel(writer, sheet_name="Tower Counts", startrow=current_row, index=False)
            current_row += 2  # Space for header and blank row
            
            # Write table data
            df.to_excel(writer, sheet_name="Tower Counts", startrow=current_row, index=False)
            
            # Apply yellow fill to column headings
            worksheet = writer.sheets["Tower Counts"]
            header_row = current_row + 1  # pandas writes headers at startrow + 1 (1-based indexing in openpyxl)
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=header_row, column=col_idx)
                cell.fill = yellow_fill
            
            current_row += len(df) + 2  # Space for table and gap between tables
    
    output.seek(0)
    return output

files = get_cos_files()    

for file in files:
    if "Veridia/Veridia Anti. Slab Cycle With Possesion dates-(15-05-2025)" in file:
        response = st.session_state.cos_client.get_object(Bucket="projectreportnew", Key=file)
     
# uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

try:
    if response:
        try:
            # Load Excel workbook
            
            wb = load_workbook(BytesIO(response['Body'].read()), data_only=True)
            sheet_names = wb.sheetnames
            # sheet_name = st.selectbox("Select a sheet", sheet_names)
            sheet_name = "Revised baseline with 60d NGT"
            
            # Get unique years from the data
            available_years = get_unique_years(wb, sheet_name)
            selected_year = st.sidebar.selectbox(
                "Select Year",
                options=available_years,
                index=len(available_years) - 1 if available_years else 0  # Default to the latest year
            )
            
            # Clear previous tables
            all_tower_tables.clear()
            
            # Check if at least one month is selected
            if selected_month_nums:
                # Process each tower
                Tower2(wb, sheet_name, selected_month_nums, selected_year)
                st.divider()
                Tower3(wb, sheet_name, selected_month_nums, selected_year)
                st.divider()
                Tower4(wb, sheet_name, selected_month_nums, selected_year)
                st.divider()
                Tower5(wb, sheet_name, selected_month_nums, selected_year)
                st.divider()
                Tower6(wb, sheet_name, selected_month_nums, selected_year)
                st.divider()
                Tower7(wb, sheet_name, selected_month_nums, selected_year)

                st.session_state.slabdf = all_tower_tables
                
                month_set = set()
                for tower in all_tower_tables:
                    for entry in tower["table_data"]:
                        month_set.update(k for k in entry if k not in ["Category", "Total"])
                months = sorted(month_set)

                completed_data = []
                non_completed_data = []

                for tower in all_tower_tables:
                    completed_row = {"Tower Name": tower["tower_name"]}
                    non_completed_row = {"Tower Name": tower["tower_name"]}
                    
                    # Initialize values
                    for month in months:
                        completed_row[month] = 0
                        non_completed_row[month] = 0

                    completed_total = 0
                    non_completed_total = 0

                    for entry in tower["table_data"]:
                        if entry["Category"] == "Completed":
                            for month in months:
                                completed_row[month] += entry.get(month, 0)
                            completed_total += int(entry.get("Total", 0))
                        elif entry["Category"] == "Non-Completed":
                            for month in months:
                                non_completed_row[month] += entry.get(month, 0)
                            non_completed_total += int(entry.get("Total", 0))

                    completed_row["Total"] = completed_total
                    non_completed_row["Total"] = non_completed_total

                    completed_data.append(completed_row)
                    non_completed_data.append(non_completed_row)

                # Convert to DataFrames
                df_completed = pd.DataFrame(completed_data)
                df_non_completed = pd.DataFrame(non_completed_data)

                # Display in Streamlit
                st.title("Tower Data - Separated by Category")

                st.subheader("✅ Completed Work")
                st.dataframe(df_completed)

                st.subheader("❌ Non-Completed Work")
                st.dataframe(df_non_completed)
            


                # Add download button if tables were generated
                if all_tower_tables:
                    st.session_state.slab = create_excel_file(all_tower_tables, selected_year)
                    st.download_button(
                        label="Download All Towers as Excel",
                        data=st.session_state.slab,
                        file_name=f"tower_counts_{selected_year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.warning("Please select at least one month to proceed.")
            
        except Exception as e:
            st.error(f"Error reading the Excel file: {e}")
except Exception as e:
    st.error(f"Slab Report File Not Found")
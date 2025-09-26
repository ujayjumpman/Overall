
import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string

towerf = []
towerg = []
towerh = []


WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [[{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()



def TowerF(sheet):
    # st.write("Analyzing Eligo Tower F")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['B', 'D', 'F', 'H']
    
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    towerf.append(1)
                if color == "FF00B0F0":
                    towerf.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")

def TowerG(sheet):
    # st.write("Analyzing Eligo Tower G")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['L', 'N', 'P', 'R', 'T', 'V']

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    towerg.append(1)
                if color == "FF00B0F0":
                    towerg.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")

def TowerH(sheet):
    # st.write("Analyzing Eligo Tower H")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['Z', 'AB', 'AD', 'AF', 'AH', 'AJ', 'AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX', 'AZ']

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    towerh.append(1)
                if color == "FF00B0F0":
                    towerh.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")



def Processjson(data):
    json_data = []
    for project, tower, green, non_green, finishing in zip(
    data["Project Name"],
    data["Tower"],
    data["Green (1)"],
    data["Non-Green (0)"],
    data["Finishing"]
):
        total = green + non_green
        structure = f"{math.ceil(green / total * 100)}%" if total > 0 else "0.00%"
        
        entry = {
            "Project": project,
            "Tower Name": tower,
            "Structure": structure,
            "Finishing": finishing
        }
        json_data.append(entry)
    
    return json_data



def ProcessGandH(exceldatas):

    wb = load_workbook(exceldatas, data_only=True)
    sheet_names = wb.sheetnames
    sheet_name = "Revised Baselines- 25 days SC"

    sheet = wb[sheet_name]
    towerf.clear()
    towerg.clear()
    towerh.clear()
    #Revised Baselines- 25 days SC
    TowerF(sheet)
    TowerG(sheet)
    TowerH(sheet)

    # st.write(towerf.count(1))
    # st.write(towerg.count(1))
    # st.write(towerh.count(1))
    data = {
    "Project Name":["ELIGO", "ELIGO", "ELIGO"],
    "Tower": ["TOWER F", "TOWER G", "TOWER H"],
    "Green (1)": [towerf.count(1), towerg.count(1), towerh.count(1)],
    "Non-Green (0)": [towerf.count(0), towerg.count(0), towerh.count(0)],
    "Finishing":[st.session_state.towerf_finishing,st.session_state.towerg_finishing,st.session_state.towerh_finishing]
}
    
     # Calculate average percentage of green
    green_counts = data["Green (1)"]
    non_green_counts = data["Non-Green (0)"]

    # Add Structure column with rounded percentages
    data["Structure"] = [f"{round(green / (green + non_green) * 100)}%" if (green + non_green) > 0 else "0%"
                        for green, non_green in zip(green_counts, non_green_counts)]
        
    # st.write("ELIGO")
    # st.dataframe(data)
    # df = pd.DataFrame(data)
    json_data = Processjson(data)
    return json_data

   
        

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
import json
from io import BytesIO


WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

t2 = []
t3 = []
t4 = []
t5 = []
t7 = []
t6 = []


def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1')
    excel_data = output.getvalue()
    return excel_data

def is_green(cell):

    fill = cell.fill
    color_code = fill.start_color.index
    cell_value = cell.value

    # st.write(cell_value) 
    # st.write(color_code) 
    if color_code == 'FF92D050' and cell_value != None or color_code == 9:
        return 1
    if color_code == "FF00B0F0":
        return 2



def GetGreenCountsTop(start_value, column_name, tower_name):
    
    anticipated_column_index = None
    for col_num, col_name in enumerate(st.session_state.datas[3], start=start_value):  
        if col_name.value == column_name:
            anticipated_column_index = col_num
            break


    if anticipated_column_index:
        green_cells_count = 0
       
        for row in range(4, 11):  
            cell = st.session_state.datas.cell(row=row, column=anticipated_column_index)  
            
            if cell.value is not None:
                if tower_name == "t2":
                    if is_green(cell) == 1: 
                        green_cells_count += 1
                        t2.append(1)
                    elif is_green(cell) == 2:
                        t2.append(0)
                elif tower_name == "t3":
                    if is_green(cell) == 1: 
                        green_cells_count += 1
                        t3.append(1)
                    elif is_green(cell) == 2:
                        t3.append(0)
                elif tower_name == "t4":
                    if is_green(cell) == 1: 
                        green_cells_count += 1
                        t4.append(1)
                    elif is_green(cell) == 2:
                        t4.append(0)
                elif tower_name == "t5":
                    if is_green(cell) == 1: 
                        green_cells_count += 1
                        t5.append(1)
                    elif is_green(cell) == 2:
                        t5.append(0)
                elif tower_name == "t6":
                    if is_green(cell) == 1: 
                        green_cells_count += 1
                        t6.append(1)
                    elif is_green(cell) == 2:
                        t6.append(0)
                elif tower_name == "t7":
                    if is_green(cell) == 1: 
                        green_cells_count += 1
                        t7.append(1)
                    elif is_green(cell) == 2:
                        t7.append(0)
                    
                
        # st.write(f"There are {green_cells_count} green-colored cells in the {column_name} column.")

def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']
    


def generatePrompt(json_datas):
    body = {
        "input": f"""
    
      Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as json but dont change the project value veridia 
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
        [{{

           "Project":"Veridia"
           "Tower Name:"tower name",
           "Structure":"percentage",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON .

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()


# results = {}
# total_entries_per_tower = {}

def Countingprocess(count_datas):
    global results, total_entries_per_tower
    try:
        # Add 'T34' column with value 'T34' for all rows
        count_datas['T34'] = 'T34'

        # Display tempdatas for verification
        st.write("### Temp DataFrame with T34 Column:")
        st.dataframe(count_datas)

        # Process columns (exclude 'T34' from results since it’s not numeric)
        columns = [col for col in count_datas.columns if col != 'T34']
        for column in columns:
            if column not in results:
                results[column] = {"0": 0, "1": 0}
            total_entries_per_tower[column] = 0  # Initialize or reset total entries
            st.write(f"Counting {column}:")
            for i in range(0, len(count_datas[column]), 10):  # chunk = 10
                datas = count_datas[column].iloc[i:i + 10]
                total_entries_per_tower[column] += len(datas)  # Track total entries
                json_datas = json.loads(generatePrompt(datas))
                if column in json_datas:
                    for key, value in json_datas[column].items():
                        results[column][key] += value

        # Update results_df with accumulated results
        st.session_state.results_df = pd.DataFrame(results)
        st.write("### Results DataFrame (Debug):")
        st.dataframe(st.session_state.results_df)

    except Exception as e:
        st.write(f"Error in Countingprocess: {e}")


def GetGreenCountsBottom(start_value, column_name, tower_name):
    # workbook = load_workbook(uploaded_file)
    # sheet = workbook['Revised Baselines']


    anticipated_column_index = None
    for col_num, col_name in enumerate(st.session_state.datas[3], start=start_value):  
        if col_name.value == column_name:
            anticipated_column_index = col_num
            break


    if anticipated_column_index:
        green_cells_count = 0
       
        for row in range(14, 21):  
            cell = st.session_state.datas.cell(row=row, column=anticipated_column_index)  
            
            if cell.value is not None:
                if tower_name == "t2":
                    if is_green(cell): 
                        green_cells_count += 1
                        t2.append(1)
                    else:
                        t2.append(0)
                elif tower_name == "t3":
                    if is_green(cell): 
                        green_cells_count += 1
                        t3.append(1)
                    else:
                        t3.append(0)
                elif tower_name == "t4":
                    if is_green(cell): 
                        green_cells_count += 1
                        t4.append(1)
                    else:
                        t4.append(0)
                elif tower_name == "t5":
                    if is_green(cell): 
                        green_cells_count += 1
                        t5.append(1)
                    else:
                        t5.append(0)
                elif tower_name == "t6":
                    if is_green(cell): 
                        green_cells_count += 1
                        t6.append(1)
                    else:
                        t6.append(0)
                elif tower_name == "t7":
                    if is_green(cell): 
                        green_cells_count += 1
                        t7.append(1)
                    else:
                        t7.append(0)
                    
        # st.write(f"There are {green_cells_count} green-colored cells in the {column_name} column.")
                




def CountingProcess3(uploaded_file):
    if uploaded_file is not None:
    # datas = pd.read_excel(uploaded_file, skiprows=0, sheet_name='Revised Baselines')

    # pour1 = datas.iloc[:, :20]

    # pour1_subset = pour1.iloc[2:17]

    # st.write("Here is the data from your uploaded Excel file (after applying your logic):")
    # st.dataframe(pour1_subset)

        workbook = load_workbook(uploaded_file)
        sheet = workbook['Revised Baselines']
        st.session_state.datas = sheet


        #1, 3, 5, 7, 9, 11, 13, 15

        tower2 = [1, 3, 5, 7, 9, 11, 13, 15]
        tower3 = [19, 21, 23, 25, 27, 29, 31, 33]
        tower4 = [37, 39, 41, 43, 45, 47, 49, 51, 53, 55, 57, 59, 61, 63, 65, 67]
        tower5 = [102,104,106, 108, 110, 112, 114, 116, 118, 120, 122, 124, 126, 128]
        tower7 = [131, 133, 135, 137, 139, 141, 143, 145, 147, 149, 151, 153, 155, 157]
        tower6 = [160, 162, 164, 166, 168, 170, 172, 174, 176, 178, 180, 182, 184, 186]


        #tower2
        st.header("Counting Tower2.....")
        for i in tower2:
            GetGreenCountsTop(i, 'M1 West', "t2")
            GetGreenCountsBottom(i, "M1 West", "t2")
        
        

        #tower3
        st.header("Counting Tower3.....")
        for i in tower3:
            GetGreenCountsTop(i, 'M1 West', "t3")
            GetGreenCountsBottom(i, "M1 West", "t3")


        #tower4
        st.header("Counting Tower4.....")
        for i in tower4:
            GetGreenCountsTop(i, 'M1 West', "t4")
            GetGreenCountsBottom(i, "M1 West", "t4")


        # tower5
        st.header("Counting Tower5.....")
        for i in tower5:
            GetGreenCountsTop(i, "M3 West","t5")
            GetGreenCountsBottom(i, "M3 West", "t5")

        #tower6
        st.header("Counting Tower6.....")
        for i in tower6:
            GetGreenCountsTop(i, "M3 West","t6")
            GetGreenCountsBottom(i, "M3 West", "t6")


        #tower7
        st.header("Counting Tower7.....")
        for i in tower7:
            GetGreenCountsTop(i, "M3 West","t7")
            GetGreenCountsBottom(i, "M3 West", "t7")
        # st.text(t2.count(0))
        # st.text(t3)
        # st.text(t4)
        # st.text(t5)
        # st.text(t6)
        # st.text(t7)
        data = {
    "Tower": ["Tower2", "Tower3", "Tower4", "Tower5", "Tower6", "Tower7"],
    "Green (1)": [t2.count(1), t3.count(1), t4.count(1), t5.count(1), t6.count(1), t7.count(1)],
    "Non-Green (0)": [t2.count(0), t3.count(0), t4.count(0), t5.count(0), t6.count(0), t7.count(0)],
    "Finishing":["0%","0%",st.session_state.veridia4,st.session_state.veridia5,"0%","0%"]
}

        df = pd.DataFrame(data)

        # Display the table
        st.table(df)
        generated = generatePrompt(df)
        json_data = json.loads(generated)
       

        # global results, total_entries_per_tower
        # results = {}
        # total_entries_per_tower = {}

        # Countingprocess(pd.DataFrame({
        #     'T2':t2,
        # }))
        # Countingprocess(pd.DataFrame({
        #     'T3':t3,
        # }))
        # Countingprocess(pd.DataFrame({
        #     'T4':t4,
        # }))
        # Countingprocess(pd.DataFrame({
        #     'T5':t5,
        # }))
        # Countingprocess(pd.DataFrame({
        #     'T':t6,
        # }))
        # Countingprocess(pd.DataFrame({
        #     'T':t7,
        # }))

        # st.write(st.session_state.results_df)

        # average_dict = {}
        # if 'results_df' in st.session_state and not st.session_state.results_df.empty:
        #     for i in st.session_state.results_df.columns:
        #         successes = st.session_state.results_df[i]["0"]  # Number of successes
        #         total = st.session_state.results_df[i]["1"]      # Total attempts
        #         if total == 0:
        #             percentage = 0  # Avoid division by zero
        #             average_dict[i] = [percentage]
        #         else:
        #             # Success rate as (successes / total * 100), capped at 100%
        #             percentage = str(round(min(100, (successes / total) * 100))) + "%"  
        #             average_dict[i] = [percentage]
        # else:
        #     st.write("No results data available.")
        #     return None

        # average_df = pd.DataFrame(average_dict).T
        # average_df = average_df.reset_index()
        # average_df.columns = ['Tower Name', 'Structure']

        # # Ensure 'Tower Name' contains only valid tower names
        # valid_tower_names = ["T2", "T3", "T4", "T5", "T6", "T7"]
        # average_df = average_df[average_df['Tower Name'].isin(valid_tower_names)]

        # # Add 'T34' column with 'T34' and reorder to make it the first column
        # average_df.insert(0, 'Project', 'Veridia')
        # average_df.insert(len(average_df.columns), 'Finishing', '0%')


        return json_data


# excel_file = to_excel(average_df)
# st.download_button(
#     label="Download as Excel",
#     data=excel_file,
#     file_name="tower_percentages.xlsx",
#     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# )
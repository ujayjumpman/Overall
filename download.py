import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string


COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"

cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)


def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return ["Error fetching COS files"]
    

files = get_cos_files()
st.write(files)

def download_file(file_key):
    try:
        # Fetch the file from COS
        response = cos_client.get_object(Bucket="projectreportnew", Key=file_key)
        file_data = response['Body'].read()
        return file_data
    except Exception as e:
        st.error(f"Error downloading file {file_key}: {e}")
        return None

if files:
    # Dropdown to select a file
    selected_file = st.selectbox("Select an Excel file to download:", files)
    
    if selected_file:
        # Download button for the selected file
        file_data = download_file(selected_file)
        if file_data:
            st.download_button(
                label=f"Download {selected_file}",
                data=io.BytesIO(file_data),
                file_name=selected_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.write("No files available to download.")
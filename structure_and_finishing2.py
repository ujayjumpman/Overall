import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
import json
from io import BytesIO


towerf = []
towerg = []
towerh = []


WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"


def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1')
    excel_data = output.getvalue()
    return excel_data


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']
    


def generatePrompt(json_datas):
    body = {
        "input": f"""
    
      Read all data from this table carefully:
         
        {json_datas}.
        
       need a average value as percentage for green as json but dont change the project value

       Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
        [{{
        
           "Project":"Eligo",
           "Tower Name":"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]
        Return the result strictly as a JSON object—no code, no explanations, only the JSON .

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()


def is_green(cell):

    fill = cell.fill
    color_code = fill.start_color.index
    cell_value = cell.value

    # st.write(cell_value) 
    # st.write(color_code) 
    if color_code == 'FF92D050':
        return True
    if color_code == 'FF00B0F0':
        return False

def GetGreenCountsBottom(start_value, column_name, tower_name):
    # workbook = load_workbook(uploaded_file)
    # sheet = workbook['Revised Baselines- 25 days SC']
    # st.write(sheet)


    anticipated_column_index = None
    for col_num, col_name in enumerate(st.session_state.datas[4], start=start_value):  
        if col_name.value == column_name:
            anticipated_column_index = col_num
            break


    if anticipated_column_index:
        green_cells_count = 0
       
        for row in range(5, 13):  
            cell = st.session_state.datas.cell(row=row, column=anticipated_column_index)  
            if cell.value is not None:
                if tower_name == "f":
                    if is_green(cell): 
                        green_cells_count += 1
                        towerf.append(1)
                    else:
                        towerf.append(0)
                elif tower_name == "g" :
                    if is_green(cell): 
                        green_cells_count += 1
                        towerg.append(1)
                    else:
                        towerg.append(0)
                else:
                    if is_green(cell): 
                        green_cells_count += 1
                        towerh.append(1)
                    else:
                        towerh.append(0)
                          

        # st.text(values)

        # st.write(f"There are {green_cells_count} green-colored cells in the {column_name} column.")
    # else:
    #     st.write("The 'Anticipated' column could not be found.")

results = {}

def Countingprocess(count_datas):
    global results
    try:
        # Add 'Eligo' column with value 'Eligo' for all rows
        count_datas['Eligo'] = 'Eligo'

        # Display tempdatas for verification
        # st.write("### Temp DataFrame with Eligo Column:")
        # st.dataframe(count_datas)

        # Process columns (exclude 'Eligo' from results since it’s not numeric)
        columns = [col for col in count_datas.columns if col != 'Eligo']
        for column in columns:
            if column not in results:
                results[column] = {"0": 0, "1": 0}
            st.write(f"Counting {column}:")
            for i in range(0, len(count_datas[column]), 10):  # chunk = 10
                datas = count_datas[column].iloc[i:i + 10]
                json_datas = json.loads(generatePrompt(datas))
                if column in json_datas:
                    for key, value in json_datas[column].items():
                        results[column][key] += value

        # Update results_df with accumulated results
        st.session_state.results_df = pd.DataFrame(results)
        # st.write("### Results DataFrame (Debug):")
        # st.dataframe(st.session_state.results_df)

    except Exception as e:
        st.write(f"Error in Countingprocess: {e}")

# uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

def CountingProcess2(uploaded_file):
    try:
        datas = pd.read_excel(uploaded_file, skiprows=1, sheet_name='Revised Baselines- 25 days SC')
        pour1 = datas.iloc[:, :20]
        pour1_subset = pour1.iloc[2:10]

        st.write("Here is the data from your uploaded Excel file (after applying your logic):")
        # st.dataframe(pour1_subset)

        workbook = load_workbook(uploaded_file)
        sheet = workbook['Revised Baselines- 25 days SC']
        st.session_state.datas = sheet

        # Clear global lists before populating
        global towerf, towerg, towerh
        towerf = []
        towerg = []
        towerh = []

        tower1 = [1, 5]
        st.write("Counting Tower F")
        for i in tower1:
            GetGreenCountsBottom(i, "Anticipated", "f")
        # st.write(towerf)

        st.write("Counting Tower G")
        tower2 = [11, 15, 19]
        for i in tower2:
            GetGreenCountsBottom(i, "Anticipated", "g")
        # st.write(towerg)

        st.write("Counting Tower H")
        tower3 = [25, 29, 33, 37, 41, 45, 46]
        for i in tower3:
            GetGreenCountsBottom(i, "Anticipated", "h")
        # st.write(towerh)

        data = {
    "Tower": ["Tower F", "Tower G", "Tower H"],
    "Green (1)": [towerf.count(1), towerg.count(1), towerh.count(1)],
    "Non-Green (0)": [towerf.count(0),towerg.count(0),towerh.count(0)],
    "Finishing":["0%",st.session_state.eligotg, st.session_state.eligoth]
}

        df = pd.DataFrame(data)

        # Display the table
        st.table(df)
        generated = generatePrompt(df)
        st.write(generated)
        json_datas = json.loads(generated)
        
        # # Call Countingprocess for each tower with accumulated results
        # Countingprocess(pd.DataFrame({"T(F)": towerf}))
        # Countingprocess(pd.DataFrame({"T(G)": towerg}))
        # Countingprocess(pd.DataFrame({"T(H)": towerh}))

        # # Calculate averages based on accumulated results
        # average_dict = {}
        # if 'results_df' in st.session_state and not st.session_state.results_df.empty:
        #     for i in st.session_state.results_df.columns:
        #         successes = st.session_state.results_df[i]["0"]  # Number of successes
        #         total = st.session_state.results_df[i]["1"]      # Total attempts
        #         if total == 0:
        #             percentage = 0  # Avoid division by zero
        #             average_dict[i] = [percentage]
        #         else:
                    
        #             percentage = str(round(min(100, (successes / total) * 100))) + "%"  
        #             average_dict[i] = [percentage]
        # else:
        #     st.write("No results data available.")
        #     return None

        # average_df = pd.DataFrame(average_dict).T
        # average_df = average_df.reset_index()
        # average_df.columns = ['Tower Name', 'Structure']

        # # Ensure 'Tower Name' contains only valid tower names
        # valid_tower_names = ["T(F)", "T(G)", "T(H)"]
        # average_df = average_df[average_df['Tower Name'].isin(valid_tower_names)]

        # # Add 'Eligo' column with 'Eligo' and reorder to make it the first column
        # average_df.insert(0, 'Project', 'Eligo')
        # average_df.insert(len(average_df.columns), 'Finishing', '0%')

        return json_datas

    except Exception as e:
        st.write(f"Error in CountingProcess2: {e}")
        return None
# if st.button("Manual count"):
#     st.write(f"In Tower F Total 1:{towerf.count(1)} Total zeros:{towerf.count(0)}")
#     st.write(f"In Tower G Total 1:{towerg.count(1)} Total zeros:{towerg.count(0)}")
#     st.write(f"In Tower H Total 1:{towerh.count(1)} Total zeros:{towerh.count(0)}")

# if st.button("Count by AI"):
    
#     st.table(st.session_state.results_df)
    # st.write(generatePrompt(towerf))
    # st.write(generatePrompt(towerg))
    # st.write(generatePrompt(towerh))


# st.download_button(
#     label="Download as Excel",
#     data=excel_file,
#     file_name="tower_percentages.xlsx",
#     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# )
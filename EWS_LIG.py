import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io


ews1 = []
ews2 = []
ews3 = []
lig1 = []
lig2 = []
lig3 = []

WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()

# def Checkcolor(sheet, )

def EWS1(sheet):
    # st.write("Analyzing Ews Tower 1")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set

            # st.write(value)
            # st.write(f"row:{ row}, col:{col}")
            # st.write(bg_color)
            # st.divider()
            if bg_color == "#92D050":
                ews1.append(1)
            if bg_color == "#0070C0":
                ews1.append(0)

def EWS2(sheet):
    # st.write("Analyzing Ews Tower 2")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['S', 'U', 'W', 'Y', 'AA', 'AC', 'AE', 'AG']
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set
    
            if bg_color == "#92D050":
                ews2.append(1)
            if bg_color == "#0070C0":
                ews2.append(0)

def EWS3(sheet):
    # st.write("Analyzing Ews Tower 3")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['AJ', 'AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX']
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set
    
            if bg_color == "#92D050":
                ews3.append(1)
            if bg_color == "#0070C0":
                ews3.append(0)


def LIG1(sheet):
    # st.write("Analyzing Lig Tower 1")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['AJ', 'AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX']
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set
           
            
            if bg_color == "#92D050":
                lig1.append(1)
            if bg_color == "#0070C0":
                lig1.append(0)
    

def LIG2(sheet):
    # st.write("Analyzing Lig Tower 2")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['S', 'U', 'W', 'Y', 'AA', 'AC', 'AE', 'AG']
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set

            if bg_color == "#92D050":
                lig2.append(1)
            if bg_color == "#0070C0":
                lig2.append(0)
    

def LIG3(sheet):
    # st.write("Analyzing Lig Tower 3")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    for col in cols:
        
        for row in rows:
            cell = sheet[f"{col}{row}"]
            value = cell.value if cell.value is not None else ""
        
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"  # Extract the last 6 chars for hex color
            else:
                bg_color = "#FFFFFF"  # Default to white if no background color is set

            if bg_color == "#92D050":
                lig3.append(1)
            if bg_color == "#0070C0":
                lig3.append(0)

def Processjson(data):
    json_data = []
    for project, tower, green, non_green, finishing in zip(
    data["Project Name"],
    data["Tower"],
    data["Green (1)"],
    data["Non-Green (0)"],
    data["Finishing"]
):
        total = green + non_green
        structure = f"{math.ceil(green / total * 100)}%" if total > 0 else "0%"
        
        entry = {
            "Project": project,
            "Tower Name": tower,
            "Structure": structure,
            "Finishing": finishing
        }
        json_data.append(entry)
    
    return json_data



def ProcessEWSLIG(exceldatas):

    wb = load_workbook(exceldatas, data_only=True)
    sheet_names = wb.sheetnames
    sheet_name = "Revised Baseline 45daysNGT+Rai"

    sheet = wb[sheet_name]

    ews1.clear()
    ews2.clear()
    ews3.clear()
    lig1.clear()
    lig2.clear()
    lig3.clear()
   
    EWS1(sheet)
   
    EWS2(sheet)
    
    EWS3(sheet)
    
    LIG1(sheet)
   
    LIG2(sheet)
   
    LIG3(sheet)

    data = {
    "Project Name": ["EWS", "EWS", "EWS", "LIG", "LIG", "LIG"],
    "Tower": ["EWST1", "EWST2", "EWST3", "LIGT1", "LIGT2", "LIGT3"],
    "Green (1)": [ews1.count(1), ews2.count(1), ews3.count(1), lig1.count(1), lig2.count(1), lig3.count(1)],
    "Non-Green (0)": [ews1.count(0), ews2.count(0), ews3.count(0), lig1.count(0), lig2.count(0), lig3.count(0)],
    "Finishing": ["0%","0%","0%","0%","0%","0%"]
}

    # Calculate average percentage of green
    green_counts = data["Green (1)"]
    non_green_counts = data["Non-Green (0)"]
    averages = []

    for green, non_green in zip(green_counts, non_green_counts):
        total = green + non_green
        avg = (green / total) * 100 if total > 0 else 0  # avoids division by zero
        averages.append(avg)

    data["Structure"] = data["Average (%)"] = [f"{(green / (green + non_green) * 100):.2f}%" if (green + non_green) > 0 else "0.00%"
                       for green, non_green in zip(green_counts, non_green_counts)]
    
    json_data = Processjson(data)
    # st.table(data)
    # df = pd.DataFrame(data)
    # st.write("EWS LIG")
    # st.dataframe(data)
    # Ai_answer = generatePrompt(df)

    # json_data = json.loads(Ai_answer)

    # st.write(json_data)
    return json_data

    # st.write(ews1.count(1))

           

       

   

    


    
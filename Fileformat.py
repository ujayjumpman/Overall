import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io



#   wb = load_workbook(exceldatas, data_only=True)
#         sheet_names = wb.sheetnames
#         for i in sheet_names:
#             st.write(i)

def CheckFile(uploaded_files):
    try:
        if uploaded_files:
            st.write("**Uploaded File Names and Their Sheets:**")
            for file in uploaded_files:
                # Display file name
                st.write(f"- {file.name}")
                # Validate if it's an Excel file
                if file.name.endswith(('.xlsx', '.xls')):
                    try:
                        # Read sheet names using pandas
                        xl = pd.ExcelFile(file)
                        sheet_names = xl.sheet_names
                        if sheet_names:
                            st.write("  Sheets:")
                            for sheet in sheet_names:
                                if sheet == "Revised Baseline 45daysNGT+Rai":
                                    st.write("EWS LIG")
                                # st.write(f"    - {sheet}")
                        else:
                            st.write("  No sheets found in this file.")
                    except Exception as e:
                        st.warning(f"  Could not read sheets for {file.name}: {str(e)}")
                else:
                    st.warning(f"  {file.name} is not a valid Excel file")
        else:
            st.write("No files uploaded. Please upload Excel files.")
    except Exception as e:
        st.error("Please upload a valid Excel file")
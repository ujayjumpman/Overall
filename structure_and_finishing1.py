import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
import json
from io import BytesIO

WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"
total_counts = {}
ewst1 = []
ewst2 = []
ewst3 = []

ligt1 = []
ligt2 = []
ligt3 = []

def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1')
    excel_data = output.getvalue()
    return excel_data


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']
    


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()




def is_green(cell):

    fill = cell.fill
    color_code = fill.start_color.index
    cell_value = cell.value

    # st.write(cell_value) 
    # st.write(color_code) 
    if color_code == 'FF92D050':
        return 1
    if color_code == 4:
        return 2
    else:
        return None
    
def GetGreenCountsTop(start_value, column_name, tower_name):
    # workbook = load_workbook(uploaded_file)
    # sheet = workbook['Revised Baseline 45daysNGT+Rai']


    anticipated_column_index = None
    for col_num, col_name in enumerate(st.session_state.datas[4], start=start_value):  
        if col_name.value == column_name:
            anticipated_column_index = col_num
            break


    if anticipated_column_index:
        green_cells_count = 0
       
        for row in range(7, 22):  
            cell = st.session_state.datas.cell(row=row, column=anticipated_column_index)  
            if tower_name == "Tower1":
                if is_green(cell) == 1: 
                    green_cells_count += 1
                    ewst1.append(1)
                elif is_green(cell) == 2:
                    ewst1.append(0)
            elif tower_name == "Tower2" :
                # st.write(is_green(cell))
                if is_green(cell) == 1: 
                    green_cells_count += 1
                    ewst2.append(1)
                elif is_green(cell) == 2:
                    ewst2.append(0)
            else:
                if is_green(cell) == 1: 
                    green_cells_count += 1
                    ewst3.append(1)
                elif is_green(cell) == 2:
                    ewst3.append(0)

def GetGreenCountsBottom(start_value, column_name, tower_name):
    


    anticipated_column_index = None
    for col_num, col_name in enumerate(st.session_state.datas[4], start=start_value):  
        if col_name.value == column_name:
            anticipated_column_index = col_num
            break


    if anticipated_column_index:
        green_cells_count = 0
       
        for row in range(29, 44):  
            cell = st.session_state.datas.cell(row=row, column=anticipated_column_index)  
            if tower_name == "Tower1":
                if is_green(cell) == 1: 
                    green_cells_count += 1
                    ligt1.append(1)
                elif is_green(cell) == 2:
                    ligt1.append(0)
            elif tower_name == "Tower2" :
                if is_green(cell) ==  1: 
                    green_cells_count += 1
                    ligt2.append(1)
                elif is_green(cell) == 2:
                    ligt2.append(0)
            else:
                if is_green(cell) == 1: 
                    green_cells_count += 1
                    ligt3.append(1)
                elif is_green(cell) == 2:
                    ligt3.append(0)


def count_zeros_ones(df):
    counts = {}
    for column in df.columns:
        counts[column] = df[column].value_counts().to_dict()
    return counts



        # st.text(values)

    #     st.write(f"There are {green_cells_count} green-colored cells in the {column_name} column.")
    # else:
    #     st.write("The 'Anticipated' column could not be found.")

    #{"tower_1": {"0": 2, "1": 3}, "tower_2": {"0": 4, "1": 1}, "tower_3": {"0": 1, "1": 4}


results = {}
chunk = 10
def CountingProcess(uploaded_file):
    try:
        datas = pd.read_excel(uploaded_file, skiprows=3, sheet_name='Revised Baseline 45daysNGT+Rai')

        pour1 = datas.iloc[:, :20]

        pour1_subset = pour1.iloc[2:17]

            # st.write("Here is the data from your uploaded Excel file (after applying your logic):")
            # st.dataframe(pour1_subset)

        workbook = load_workbook(uploaded_file)
        sheet = workbook['Revised Baseline 45daysNGT+Rai']
        st.session_state.datas = sheet

        # GetGreenCountsTop(1,"Anticipated", "Tower1")


        
        tower1 = [1,5,9,13]
        # st.write("Counting EWS Tower1......")
        for i in tower1:
            GetGreenCountsTop(i, "Anticipated", "Tower1")
        

        tower2 = [20,24,28,32]

        # st.write("Counting EWS Tower2......")
        for i in  tower2:
            GetGreenCountsTop(i, "Baseline","Tower2")


        tower3 = [37,41,45,49]


        # st.write("Counting EWS Tower3......")
        for i in  tower3:
            GetGreenCountsTop(i, "Baseline", "Tower3")


        # st.write("Counting LIG Tower1......")
        for i in tower1:
            GetGreenCountsBottom(i, "Anticipated", "Tower3")

        # st.write("Counting LIG Tower2......")
        for i in tower2:
            GetGreenCountsBottom(i, "Baseline", "Tower2")

        # st.write("Counting LIG Tower3......")
        for i in tower3:
            GetGreenCountsBottom(i, "Baseline", "Tower1")
        # st.header("Tower 1")
        # st.write(ewst1)

        data = {
    "Project Name":["EWS", "EWS", "EWS", "LIG", "LIG", "LIG"],
    "Tower": ["EWST1", "EWST2", "EWST3", "LIGT1", "LIGT2", "LIGT3"],
    "Green (1)": [ewst1.count(1), ewst2.count(1), ewst3.count(1), ligt1.count(1), ligt2.count(1), ligt3.count(1)],
    "Non-Green (0)": [ewst1.count(0), ewst2.count(0), ewst3.count(0), ligt1.count(0), ligt2.count(0), ligt3.count(0)],
}
        df = pd.DataFrame(data)

        # Display the table
        st.table(df)
        g = generatePrompt(df)
        st.write(g)
        json_datas = json.loads(g)
        st.text(json_datas)
        # st.write(generatePrompt(df))
       

        # st.session_state.tempdatas = pd.DataFrame({
        #     'EWST1': ewst1,  # EWS Tower 1
        #     'EWST2': ewst2,  # EWS Tower 2
        #     'EWST3': ewst3,  # EWS Tower 3
        #     'LIGT1': ligt1,  # LIG Tower 1
        #     'LIGT2': ligt2,  # LIG Tower 2
        #     'LIGT3': ligt3,  # LIG Tower 3
        # })

        # # Add 'Project' column with value 'Eden' for all rows
        # st.session_state.tempdatas['Project'] = 'Eden'

        # # Display tempdatas for verification
        # # st.write("### Temp DataFrame with Project Column:")
        # # st.dataframe(st.session_state.tempdatas)

        # # Process columns (exclude 'Project' from results since it’s not numeric)
        # columns = [col for col in st.session_state.tempdatas.columns if col != 'Project']
        # for column in columns:
        #     results[column] = {"0": 0, "1": 0}
        #     for i in range(0, len(st.session_state.tempdatas[column]), chunk):
        #         datas = st.session_state.tempdatas[column].iloc[i:i + chunk]
        #         json_datas = json.loads(generatePrompt(datas))
        #         if column in json_datas:
        #             for key, value in json_datas[column].items():
        #                 results[column][key] += value

        # # Create results DataFrame (only for numeric columns)
        # st.session_state.results_df = pd.DataFrame(results)
        # # st.write("### Results DataFrame (Debug):")
        # # st.dataframe(st.session_state.results_df)

        
        # average_dict = {}
        # columns = [col for col in st.session_state.tempdatas.columns if col != 'Project']
        # total_length = len(st.session_state.tempdatas)  

        # for column in columns:
            
        #     successes = st.session_state.tempdatas[column].sum()  # Sum of 1s
            
        #     percentage = str(round(min(100, (successes / total_length) * 100))) + "%"  
        #     average_dict[column] = [percentage] 

        # # Create average_df with tower data
        # average_df = pd.DataFrame(average_dict).T
        # average_df = average_df.reset_index()
        # average_df.columns = ['Tower Name', 'Structure']

        # # Add 'Project' column with 'Eden' and reorder to match image
        # average_df.insert(0, 'Project', ["EWS","EWS","EWS","LIG","LIG","LIG"],allow_duplicates=True)

        # average_df.insert(len(average_df.columns), 'Finishing', '0%')

        return json_datas
        # st.table(average_df)
        # excel_file = to_excel(average_df)
        # st.download_button(
        #     label="Download as Excel",
        #     data=excel_file,
        #     file_name="tower_percentages.xlsx",
        #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # )

            
    except Exception as e:
        st.write(e)
    



# st.write(st.session_state.tempdatas)
# if st.button("Manual Count"):
#     tower_counts = count_zeros_ones(st.session_state.tempdatas)
#     st.write("Counts of 0s and 1s for each tower:")
#     for tower, count in tower_counts.items():
#         st.write(f"{tower}: 0's = {count.get(0, 0)}, 1's = {count.get(1, 0)}")

# chunk = 10
# results = {}
# if st.button("Count By AI"):
    
        
   